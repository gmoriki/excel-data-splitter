# Explain by GPT-4
"""
このコードは、Excelファイルからデータを処理し、新しいExcelファイルに書き出すプログラムです。以下に、主な機能をわかりやすく説明します。

Excelファイルを読み込む。
シート内の結合されたセルを解除し、元のセルと同じ値を入れる。
隣接するセルが空でなければ、枠線を引く。
枠線で囲まれたテーブルを見つける。
見つけたテーブルをデータフレーム（表形式のデータ）に変換する。
データフレームを新しいExcelファイルの複数シートに書き出す。
このプログラムでは、まずinput.xlsxというExcelファイルを読み込みます。次に、各シートに対して処理を行い、最後にoutput.xlsxという名前の新しいExcelファイルに書き出します。

具体的な処理の流れは以下の通りです。

結合されたセルを解除し、元のセルと同じ値を入れる。
隣接するセルが空でなければ、枠線を引く。
枠線で囲まれていないセルを削除する。
シート内の枠線で括られたテーブルを検出する。
検出したテーブルをデータフレームに変換する。
データフレームを新しいExcelファイルの複数シートに出力する。
このプログラムを実行すると、input.xlsxの内容が処理されて、output.xlsxに書き出されます。

"""

import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

import pandas as pd


def apply_border_to_adjacent_cells(ws):
    max_row = ws.max_row
    max_col = ws.max_column

    def is_adjacent_cell_empty(cell):
        row, col = cell.row, cell.column
        left = ws.cell(row=row, column=col - 1) if col > 1 else None
        right = ws.cell(row=row, column=col + 1) if col < max_col else None
        top = ws.cell(row=row - 1, column=col) if row > 1 else None
        bottom = ws.cell(row=row + 1, column=col) if row < max_row else None

        return not (left and left.value or right and right.value or top and top.value or bottom and bottom.value)

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    for row in ws.iter_rows(min_row=1, max_col=max_col, max_row=max_row):
        for cell in row:
            if cell.value and not is_adjacent_cell_empty(cell):
                cell.border = thin_border


def apply_border_to_range(ws, min_row, max_row, min_col, max_col):
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).border = thin_border


def unmerge_and_fill_cells(ws):
    merged_ranges = ws.merged_cells.ranges.copy()
    for merged_cells in merged_ranges:
        top_left_cell = ws.cell(row=merged_cells.min_row, column=merged_cells.min_col)
        top_left_value = top_left_cell.value
        ws.unmerge_cells(str(merged_cells))
        apply_border_to_range(
            ws, merged_cells.min_row, merged_cells.max_row, merged_cells.min_col, merged_cells.max_col
        )
        for row in range(merged_cells.min_row, merged_cells.max_row + 1):
            for col in range(merged_cells.min_col, merged_cells.max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is None:
                    cell.value = top_left_value

    # プレーンテキストに変換
    for i, row in enumerate(ws.iter_rows(), start=1):
        for j, cell in enumerate(row, start=1):
            ws.cell(row=i, column=j).value = cell.value


def find_bordered_tables(ws):
    tables = []
    max_row = ws.max_row
    max_col = ws.max_column
    visited_cells = set()

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if (r, c) not in visited_cells:
                cell = ws.cell(row=r, column=c)
                if cell.border.left.style and cell.border.top.style:
                    start = cell.coordinate
                    r_end, c_end = r, c

                    while ws.cell(row=r_end + 1, column=c).border.top.style:
                        r_end += 1
                    while ws.cell(row=r, column=c_end + 1).border.left.style:
                        c_end += 1

                    end = ws.cell(row=r_end, column=c_end).coordinate
                    tables.append((start, end))

                    for row in range(r, r_end + 1):
                        for col in range(c, c_end + 1):
                            visited_cells.add((row, col))
    return tables


def read_table_as_dataframe(ws, start, end):
    data = []
    for row in ws[start:end]:
        data.append([cell.value for cell in row])
    return pd.DataFrame(data[1:], columns=data[0])


def save_dataframe_to_excel_sheet(df, workbook, sheet_name):
    ws = workbook.create_sheet(sheet_name)

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)


def unmerge_and_delete_unbordered_cells(ws):
    # 枠線で囲まれたセルの範囲を特定する
    bordered_cells = []

    for row in ws:
        for cell in row:
            if any([cell.border.left.style, cell.border.right.style, cell.border.top.style, cell.border.bottom.style]):
                bordered_cells.append(cell.coordinate)

    # 枠線で囲まれていないセルの結合を解除する
    for merge_range in ws.merged_cells.ranges:
        is_outside_borders = False
        for cell in merge_range:
            if cell.coordinate not in bordered_cells:
                is_outside_borders = True
                break

        if is_outside_borders:
            ws.unmerge_cells(str(merge_range))

    # 枠線で囲まれていないセルを削除する
    for row in ws.iter_rows():
        for cell in row:
            if cell.coordinate not in bordered_cells:
                ws[cell.coordinate].value = None


def convert_formula_to_plain_text(ws):
    # 関数の返り値をプレーンテキストに変換する
    for i, row in enumerate(ws.iter_rows(), start=1):
        for j, cell in enumerate(row, start=1):
            ws.cell(row=i, column=j).value = cell.value

    # return new_worksheet


def main(input_excel_file, output_excel_file):
    # input_excel_file = "input.xlsx"
    # Excelファイルを開く
    wb = openpyxl.load_workbook(input_excel_file, data_only=True)

    # output_excel_file = "output.xlsx"
    output_wb = openpyxl.Workbook()
    output_wb.remove(output_wb.active)  # デフォルトの空シートを削除

    for ws in wb.worksheets:
        # 枠線で囲まれていないセルを全て削除
        unmerge_and_delete_unbordered_cells(ws)

        # 隣接するセルのみテーブルと定めて、枠線を引く処理
        apply_border_to_adjacent_cells(ws)

        # 結合セルを解除し、元のセルと同じ値を代入
        unmerge_and_fill_cells(ws)

        # シート内の枠線で括られたテーブルを検出
        tables = find_bordered_tables(ws)

        # テーブルを読み込み
        dataframes = [read_table_as_dataframe(ws, start, end) for start, end in tables]

        # データフレームを新しいExcelファイルの複数シートに出力
        for idx, df in enumerate(dataframes):
            sheet_name = f"{ws.title}_Table_{idx+1}"
            save_dataframe_to_excel_sheet(df, output_wb, sheet_name)

    output_wb.save(output_excel_file)


main(r"./input.xlsx", r"./output.xlsx")
